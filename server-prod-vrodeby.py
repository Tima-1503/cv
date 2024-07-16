from flask import Flask, request, jsonify
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Секретный ключ для проверки токена (замените на свой)
SECRET_KEY = 'rAO0woSmvvze8g3XKIQVmPvaGammf5MhbHMCwGagTKHgu0ZuDuyqTiRH2zV8VUfA'

# Словарь для отслеживания текущих статусов по компьютерам и пользователям
status_dict = {}

# Глобальные переменные для отслеживания временных периодов
presence_periods = []
absence_periods = []
current_presence_period = None
current_absence_period = None
total_time_present = timedelta(seconds=0)
total_time_absent = timedelta(seconds=0)
previous_date = datetime.now().strftime('%d-%m-%Y')  # Переменная для хранения предыдущей даты


# Функция для проверки токена авторизации
def verify_token(token):
    return token == SECRET_KEY  # Простая проверка на совпадение с секретным ключом


# Декоратор для проверки токена
def token_required(f):
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token or not verify_token(token.split()[-1]):
            return jsonify({'error': 'Unauthorized access'}), 401
        return f(*args, **kwargs)

    return decorated_function


# Функция для обновления статуса в словаре
def update_status(computer_name, username, status):
    global current_presence_period, current_absence_period, total_time_present, total_time_absent, presence_periods, absence_periods

    if computer_name not in status_dict:
        status_dict[computer_name] = {}
    if username not in status_dict[computer_name]:
        status_dict[computer_name][username] = {'status': False, 'last_change_time': None}

    current_status = status_dict[computer_name][username]['status']
    if current_status != status:
        status_dict[computer_name][username]['status'] = status
        status_dict[computer_name][username]['last_change_time'] = datetime.now()

        # Получаем текущее время для записи периодов времени присутствия и отсутствия
        current_time = datetime.now()

        if status:  # Присутствие лица
            if current_presence_period is None:
                current_presence_period = f"{current_time.strftime('%H:%M:%S')} - "
            if current_absence_period is not None:
                end_time_absent = current_time.strftime('%H:%M:%S')
                absence_periods.append(f"{current_absence_period}{end_time_absent}")
                absence_duration = datetime.strptime(end_time_absent, '%H:%M:%S') - datetime.strptime(
                    current_absence_period.split(" - ")[0], '%H:%M:%S')
                total_time_absent += absence_duration
                current_absence_period = None

        else:  # Отсутствие лица
            if current_absence_period is None:
                current_absence_period = f"{current_time.strftime('%H:%M:%S')} - "
            if current_presence_period is not None:
                end_time_present = current_time.strftime('%H:%M:%S')
                presence_periods.append(f"{current_presence_period}{end_time_present}")
                presence_duration = datetime.strptime(end_time_present, '%H:%M:%S') - datetime.strptime(
                    current_presence_period.split(" - ")[0], '%H:%M:%S')
                total_time_present += presence_duration
                current_presence_period = None


# Функция для записи данных в файл Xlsx
def write_to_excel(computer_name, username, current_date, presence_periods, absence_periods, total_work_time,
                   total_absence_time):
    filename = f"{computer_name}-{username}-{current_date}.xlsx"
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    # подпись столбцов
    ws['A1']= 'Имя пк'
    ws['B1'] = 'Имя пользователя'
    ws['C1'] = 'Дата'
    ws['D1'] = 'Период времени Присутствия'
    ws['E1'] = 'Период времени отсутствия'
    ws['F1']= 'общее время Работы'
    ws['G1'] = 'Общее время отсутствия'
          # Запись данных в соответствующие столбцы
    row = [computer_name, username, current_date, presence_periods[-1] if presence_periods else "",
           absence_periods[-1] if absence_periods else "", total_work_time, total_absence_time]
    ws.append(row)

    wb.save(filename)


# Роут для приема данных от клиента
@app.route('/', methods=['POST'])
@token_required
def receive_data():
    global current_presence_period, current_absence_period, total_time_present, total_time_absent, presence_periods, absence_periods, previous_date

    data = request.json
    computer_name = data.get('computer_name')
    username = data.get('username')
    status = data.get('status')
    current_date = datetime.now().strftime('%d-%m-%Y')

    # Проверка, если текущая дата отличается от предыдущей
    if current_date != previous_date:
        # Сброс глобальных переменных для нового дня
        presence_periods = []
        absence_periods = []
        current_presence_period = None
        current_absence_period = None
        total_time_present = timedelta(seconds=0)
        total_time_absent = timedelta(seconds=0)
        previous_date = current_date  # Обновляем предыдущую дату

    if computer_name and username and status is not None:
        update_status(computer_name, username, status)

        # Вычисление общего времени работы и отсутствия
        total_work_time = total_time_present.total_seconds() if total_time_present else 0
        total_absence_time = total_time_absent.total_seconds() if total_time_absent else 0

        # Запись данных в Xlsx файл
        write_to_excel(computer_name, username, current_date, presence_periods, absence_periods, total_work_time,
                       total_absence_time)

        return jsonify({'message': 'Data received and processed successfully'}), 200
    else:
        return jsonify({'error': 'Invalid data format'}), 400


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
''